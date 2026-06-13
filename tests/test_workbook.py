from __future__ import annotations

import hashlib
import os
import shutil
import subprocess
import sys
import zipfile
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from xml.etree import ElementTree

import openpyxl
import pytest
from openpyxl.utils import get_column_letter

from src.tax_model import (
    add_months,
    annual_tax,
    event_tax_batches,
    installment_status,
    option_income,
    restricted_stock_income,
)


ROOT = Path(__file__).resolve().parents[1]
BUILDER = ROOT / "src" / "build_workbook.py"
OUTPUT = ROOT / "广电计量股权激励个人所得税计算及分期台账-2026版.xlsx"
ORIGINAL = ROOT / "股权激励个人所得税计算表-V2-2026.06.11 17点.xlsx"
ORIGINAL_SHA256 = "A320E140E68AF70EAC909E90502C5BC120882D447CFEB141415AA5FE0063E6A0"
SHEETS = [
    "使用说明",
    "计划参数",
    "激励事件明细",
    "年度计税汇总",
    "分期缴税台账",
    "单人测算",
    "税率及规则",
    "政策及公告",
]
ERROR_TOKENS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?")
XML_NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest().upper()


def excel_com_open(path: Path, save: bool = False) -> subprocess.CompletedProcess[str]:
    script = r"""
[Console]::OutputEncoding=[System.Text.Encoding]::UTF8
$excel=$null; $wb=$null
try {
  $excel=New-Object -ComObject Excel.Application
  $excel.Visible=$false
  $excel.DisplayAlerts=$false
  $wb=$excel.Workbooks.Open($env:XLSX,0,$false)
  $excel.CalculateFullRebuild()
  if ($env:SAVE -eq '1') { $wb.Save() }
  Write-Output ("COM_OPEN_OK Excel=" + $excel.Version + " Sheets=" + $wb.Worksheets.Count)
  $wb.Close($false)
} catch {
  Write-Output ("COM_OPEN_ERROR=" + $_.Exception.Message)
  Write-Output ("HRESULT=0x{0:X8}" -f ($_.Exception.HResult -band 0xffffffff))
  exit 1
} finally {
  if ($wb) { [void][Runtime.InteropServices.Marshal]::ReleaseComObject($wb) }
  if ($excel) {
    $excel.Quit()
    [void][Runtime.InteropServices.Marshal]::ReleaseComObject($excel)
  }
  [GC]::Collect()
  [GC]::WaitForPendingFinalizers()
}
"""
    env = os.environ.copy()
    env["XLSX"] = str(path)
    env["SAVE"] = "1" if save else "0"
    return subprocess.run(
        ["powershell", "-NoProfile", "-Command", script],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        env=env,
        timeout=120,
    )


def validate_xlsx_xml(path: Path) -> None:
    with zipfile.ZipFile(path) as archive:
        assert archive.testzip() is None
        for name in archive.namelist():
            if name.endswith((".xml", ".rels")):
                ElementTree.fromstring(archive.read(name))


def test_committed_artifact_is_valid_xml_and_excel_can_open():
    validate_xlsx_xml(OUTPUT)
    if sys.platform != "win32":
        pytest.skip("Excel COM仅适用于Windows")
    result = excel_com_open(OUTPUT)
    assert result.returncode == 0, result.stdout + result.stderr
    assert "COM_OPEN_OK Excel=16.0" in result.stdout


@pytest.fixture(scope="session")
def generated_path(tmp_path_factory):
    assert BUILDER.exists(), "工作簿生成器尚未创建"
    path = tmp_path_factory.mktemp("workbook") / OUTPUT.name
    env = os.environ.copy()
    env["WORKBOOK_OUTPUT"] = str(path)
    subprocess.run([sys.executable, str(BUILDER)], cwd=ROOT, env=env, check=True)
    assert path.exists(), "临时新版工作簿尚未生成"
    return path


@pytest.fixture(scope="session")
def workbook(generated_path):
    return openpyxl.load_workbook(generated_path, data_only=False)


def test_generated_artifact_is_valid_xml_and_excel_can_open(generated_path):
    validate_xlsx_xml(generated_path)
    if sys.platform != "win32":
        pytest.skip("Excel COM仅适用于Windows")
    result = excel_com_open(generated_path)
    assert result.returncode == 0, result.stdout + result.stderr


@pytest.fixture(scope="session")
def recalculated_path(generated_path, tmp_path_factory):
    path = tmp_path_factory.mktemp("recalculated") / OUTPUT.name
    shutil.copy2(generated_path, path)
    result = excel_com_open(path, save=True)
    assert result.returncode == 0, result.stdout + result.stderr
    return path


def header_map(ws) -> dict[str, int]:
    return {cell.value: cell.column for cell in ws[1] if cell.value}


def formula_cells(ws):
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == "f":
                yield cell


def test_original_workbook_is_unchanged():
    assert sha256(ORIGINAL) == ORIGINAL_SHA256


def test_workbook_has_exactly_eight_visible_business_sheets(workbook):
    assert workbook.sheetnames == SHEETS
    assert all(workbook[name].sheet_state == "visible" for name in SHEETS)
    assert workbook.calculation.calcMode == "auto"


def test_plan_parameters_include_verified_announcement_facts(workbook):
    ws = workbook["计划参数"]
    values = [cell.value for row in ws.iter_rows() for cell in row if cell.value is not None]
    text = "\n".join(str(value) for value in values)
    assert "2023年股票期权与限制性股票激励计划" in text
    assert "实际授予" in text and "完成登记" in text
    normalized_dates = {
        value.date() if isinstance(value, datetime) else value
        for value in values
        if isinstance(value, (date, datetime))
    }
    assert date(2024, 8, 27) in normalized_dates
    assert 13.06 in values
    for price in (14.56, 14.32, 14.18, 14.04):
        assert price in values
    assert "802万股不得作为个人分母" in text
    assert ws.protection.sheet
    assert ws["I1"].value == "员工编号"
    assert ws["J1"].value == "姓名"
    assert ws["I2"].value == "DEMO001"
    assert not ws["I2"].protection.locked
    assert not ws["J2"].protection.locked


def test_event_sheet_has_required_columns_and_500_input_rows(workbook):
    ws = workbook["激励事件明细"]
    headers = header_map(ws)
    required = {
        "事件编号",
        "税额批次",
        "员工编号",
        "姓名",
        "激励计划",
        "权益类型",
        "事件日期",
        "纳税年度",
        "本次行权或解禁数量",
        "行权日或解禁日收盘价",
        "适用行权或授予价格",
        "限制性股票登记日收盘价",
        "员工获授限制性股票总数",
        "员工实际出资总额",
        "原始本次所得",
        "本次应纳税所得额",
        "此前年度累计所得",
        "本次事件后年度累计所得",
        "累计税额",
        "此前累计已确认税额",
        "本次新增税额",
        "事件级36个月截止日",
        "校验",
    }
    assert required <= headers.keys()
    assert ws.max_row >= 501
    assert ws.freeze_panes == "A2"
    assert ws.tables

    formula_headers = {
        "税额批次",
        "纳税年度",
        "适用行权或授予价格",
        "限制性股票登记日收盘价",
        "原始本次所得",
        "本次应纳税所得额",
        "此前年度累计所得",
        "本次事件后年度累计所得",
        "累计税额",
        "此前累计已确认税额",
        "本次新增税额",
        "事件级36个月截止日",
        "校验",
    }
    for header in formula_headers:
        col = get_column_letter(headers[header])
        assert ws[f"{col}501"].data_type == "f", f"{header}未填充到第500条数据行"

    employee_validation = next(
        item
        for item in ws.data_validations.dataValidation
        if "C2:C501" in str(item.sqref)
    )
    assert employee_validation.formula1 == "=员工主数据编号"
    assert employee_validation.errorStyle == "information"
    assert employee_validation.showErrorMessage


def test_demo_rows_and_expected_tax_match_tax_model(workbook):
    ws = workbook["激励事件明细"]
    headers = header_map(ws)
    employee_col = headers["员工编号"]
    demo_rows = [
        row
        for row in range(2, ws.max_row + 1)
        if ws.cell(row, employee_col).value == "DEMO001"
    ]
    assert demo_rows == [2, 3]

    option_result, _ = option_income("17.80", "14.56", "50000")
    restricted_result, _ = restricted_stock_income(
        "13.06", "17.80", "50000", "50000", "434000"
    )
    assert option_result == Decimal("162000.00")
    assert restricted_result == Decimal("337500.00")
    assert annual_tax(option_result)["tax"] == Decimal("15480.00")
    assert annual_tax(option_result + restricted_result)["tax"] == Decimal("96930.00")
    batches = event_tax_batches(
        [
            {
                "event_id": "DEMO-OPT-001",
                "employee_id": "DEMO001",
                "event_date": date(2025, 1, 9),
                "income": option_result,
            },
            {
                "event_id": "DEMO-RS-001",
                "employee_id": "DEMO001",
                "event_date": date(2025, 6, 30),
                "income": restricted_result,
            },
        ]
    )
    assert [item["incremental_tax"] for item in batches] == [
        Decimal("15480.00"),
        Decimal("81450.00"),
    ]
    assert [item["deadline"] for item in batches] == [
        date(2028, 1, 9),
        date(2028, 6, 30),
    ]

    assert ws.cell(2, headers["本次行权或解禁数量"]).value == 50000
    assert ws.cell(2, headers["行权日或解禁日收盘价"]).value == 17.8
    assert ws.cell(3, headers["员工获授限制性股票总数"]).value == 50000
    assert ws.cell(3, headers["员工实际出资总额"]).value == 434000
    assert "仅演示" in ws.cell(2, headers["备注"]).value
    assert "仅演示" in ws.cell(3, headers["备注"]).value

    formulas = {
        header: ws.cell(row, headers[header]).value
        for row in demo_rows
        for header in (
            "本次应纳税所得额",
            "此前年度累计所得",
            "本次事件后年度累计所得",
            "累计税额",
            "此前累计已确认税额",
            "本次新增税额",
            "事件级36个月截止日",
        )
    }
    joined = "\n".join(formulas.values())
    assert "SUMIFS" in joined
    assert "EDATE" in joined
    assert "VLOOKUP" in joined


def test_demo_mirror_covers_cross_year_and_month_end():
    batches = event_tax_batches(
        [
            {
                "event_id": "Y2024",
                "employee_id": "E001",
                "event_date": date(2024, 12, 31),
                "income": Decimal("36000"),
            },
            {
                "event_id": "Y2025",
                "employee_id": "E001",
                "event_date": date(2025, 1, 31),
                "income": Decimal("36000"),
            },
        ]
    )
    assert [item["cumulative_income"] for item in batches] == [
        Decimal("36000.00"),
        Decimal("36000.00"),
    ]
    assert add_months(date(2024, 2, 29), 36) == date(2027, 2, 28)
    assert add_months(date(2025, 1, 31), 36) == date(2028, 1, 31)


def test_event_accumulation_uses_date_and_event_id_stable_order(workbook):
    ws = workbook["激励事件明细"]
    headers = header_map(ws)
    formula = ws.cell(2, headers["此前年度累计所得"]).value
    assert "事件日期" not in formula
    date_col = get_column_letter(headers["事件日期"])
    event_col = get_column_letter(headers["事件编号"])
    assert f"${date_col}$2:${date_col}$501" in formula
    assert f"${event_col}$2:${event_col}$501" in formula
    assert '"<"&' in formula
    assert '"<"&' in formula or '"<="' in formula


def test_installment_ledger_links_event_batch_and_uses_event_deadline(workbook):
    ws = workbook["分期缴税台账"]
    headers = header_map(ws)
    required = {
        "员工编号",
        "姓名",
        "纳税年度",
        "事件编号",
        "新增税额批次",
        "该事件新增应纳税额",
        "该事件纳税义务发生日",
        "该事件最晚缴清日",
        "缴税日期",
        "计划缴税金额",
        "实际缴税金额",
        "累计已缴",
        "剩余税额",
        "是否逾期",
        "关联有效",
        "解析事件编号",
        "解析税额批次",
        "校验",
    }
    assert required <= headers.keys()
    deadline_formula = ws.cell(2, headers["该事件最晚缴清日"]).value
    validation_formula = ws.cell(2, headers["校验"]).value
    link_formula = ws.cell(2, headers["关联有效"]).value
    paid_formula = ws.cell(2, headers["累计已缴"]).value
    departure_formula = ws.cell(2, headers["离职日期"]).value
    resolved_event_formula = ws.cell(2, headers["解析事件编号"]).value
    resolved_batch_formula = ws.cell(2, headers["解析税额批次"]).value
    assert "激励事件明细" in deadline_formula
    assert "冲突" in validation_formula
    assert "离职后缴税" in validation_formula
    assert "COUNTIF" in resolved_event_formula
    assert "有效" in link_formula
    assert get_column_letter(headers["解析事件编号"]) in paid_formula
    assert "INDEX" in resolved_event_formula and "MATCH" in resolved_event_formula
    assert "INDEX" in resolved_batch_formula and "MATCH" in resolved_batch_formula
    assert "ISNUMBER" in validation_formula
    assert '=0,""' in departure_formula
    assert ws.max_row >= 501


def test_event_and_batch_conflict_and_departure_late_payment_mirror():
    conflict = installment_status(
        "E1",
        "1000",
        [
            {
                "event_id": "E1",
                "tax_batch_id": "E2-TAX",
                "payment_date": date(2025, 2, 1),
                "amount": "1000",
            }
        ],
        date(2028, 1, 1),
        tax_batch_id="E1-TAX",
        as_of_date=date(2025, 2, 2),
    )
    assert conflict["paid"] == Decimal("0.00")
    assert conflict["has_invalid_payments"]

    late_departure = installment_status(
        "E1",
        "1000",
        [
            {
                "event_id": "E1",
                "tax_batch_id": "E1-TAX",
                "payment_date": date(2025, 3, 2),
                "amount": "1000",
            }
        ],
        date(2028, 1, 1),
        departure_date=date(2025, 3, 1),
        tax_batch_id="E1-TAX",
        as_of_date=date(2025, 3, 3),
    )
    assert late_departure["paid"] == Decimal("1000.00")
    assert late_departure["departure_unpaid"]


def test_departure_date_boundaries_after_com_recalculation(generated_path, tmp_path):
    path = tmp_path / "departure-boundaries.xlsx"
    shutil.copy2(generated_path, path)
    workbook = openpyxl.load_workbook(path)
    plan_ws = workbook["计划参数"]
    event_ws = workbook["激励事件明细"]
    ledger_ws = workbook["分期缴税台账"]
    event_headers = header_map(event_ws)
    ledger_headers = header_map(ledger_ws)

    cases = [
        ("NODEPART", "无离职员工", None, date(2025, 2, 1)),
        ("SAMEDAY", "当天缴款员工", date(2025, 2, 1), date(2025, 2, 1)),
        ("LATEPAY", "晚一天员工", date(2025, 2, 1), date(2025, 2, 2)),
    ]
    for offset, (employee_id, name, departure_date, payment_date) in enumerate(
        cases, start=3
    ):
        plan_ws.cell(offset, 9, employee_id)
        plan_ws.cell(offset, 10, name)
        event_row = offset + 1
        event_values = {
            "事件编号": f"{employee_id}-E1",
            "员工编号": employee_id,
            "姓名": name,
            "激励计划": "2023年股票期权与限制性股票激励计划",
            "权益类型": "股票期权",
            "事件日期": date(2025, 1, 9),
            "本次行权或解禁数量": 50000,
            "行权日或解禁日收盘价": 17.8,
            "离职日期": departure_date,
            "备案状态": "已备案",
        }
        for header, value in event_values.items():
            event_ws.cell(event_row, event_headers[header], value)

        ledger_row = offset - 1
        ledger_ws.cell(ledger_row, ledger_headers["事件编号"], f"{employee_id}-E1")
        ledger_ws.cell(ledger_row, ledger_headers["缴税日期"], payment_date)
        ledger_ws.cell(ledger_row, ledger_headers["实际缴税金额"], 15480)

    workbook.save(path)
    result = excel_com_open(path, save=True)
    assert result.returncode == 0, result.stdout + result.stderr

    calculated = openpyxl.load_workbook(path, data_only=True)
    ledger = calculated["分期缴税台账"]
    assert ledger.cell(2, ledger_headers["离职日期"]).value is None
    assert ledger.cell(2, ledger_headers["校验"]).value == "正常"
    assert ledger.cell(3, ledger_headers["校验"]).value == "正常"
    assert ledger.cell(4, ledger_headers["校验"]).value == "异常：离职后缴税"


def test_payment_as_of_date_formula_structure(workbook):
    instruction_ws = workbook["使用说明"]
    ledger_ws = workbook["分期缴税台账"]
    event_ws = workbook["激励事件明细"]
    summary_ws = workbook["年度计税汇总"]
    ledger_headers = header_map(ledger_ws)
    event_headers = header_map(event_ws)
    summary_headers = header_map(summary_ws)

    assert instruction_ws["A21"].value == "检查截至日期"
    assert instruction_ws["B21"].value is not None
    assert "截至日有效实缴" in ledger_headers
    effective_formula = ledger_ws.cell(2, ledger_headers["截至日有效实缴"]).value
    assert "检查截至日期" in effective_formula
    assert "ISNUMBER(I2)" in effective_formula
    assert "I2<=检查截至日期" in effective_formula
    assert f"${get_column_letter(ledger_headers['截至日有效实缴'])}$2" in event_ws.cell(
        2, event_headers["批次有效已缴"]
    ).value
    assert f"${get_column_letter(ledger_headers['截至日有效实缴'])}$2" in summary_ws.cell(
        2, summary_headers["已缴金额"]
    ).value
    assert "缴税日期为空" in ledger_ws.cell(2, ledger_headers["校验"]).value


def test_duplicate_event_formula_structure(workbook):
    event_ws = workbook["激励事件明细"]
    ledger_ws = workbook["分期缴税台账"]
    headers = header_map(event_ws)
    assert "有效事件" in headers
    valid_formula = event_ws.cell(4, headers["有效事件"]).value
    assert "COUNTIF($A$2:$A$501,A4)=1" in valid_formula
    assert headers["有效事件"] > headers["年度未缴到期键"]
    assert f"{get_column_letter(headers['有效事件'])}4" in event_ws.cell(
        4, headers["本次应纳税所得额"]
    ).value
    raw_income_formula = event_ws.cell(4, headers["原始本次所得"]).value
    assert raw_income_formula.count("(") == raw_income_formula.count(")")
    resolved_event_formula = ledger_ws.cell(
        2, header_map(ledger_ws)["解析事件编号"]
    ).value
    assert "COUNTIF('激励事件明细'!$A$2:$A$501,D2)=1" in resolved_event_formula
    assert "COUNTIF('激励事件明细'!$B$2:$B$501,E2)=1" in resolved_event_formula


def test_date_validations_use_long_term_upper_bound(workbook):
    for sheet_name in ("激励事件明细", "分期缴税台账"):
        validations = workbook[sheet_name].data_validations.dataValidation
        date_validations = [item for item in validations if item.type == "date"]
        assert date_validations
        assert all("2035" not in str(item.formula2) for item in date_validations)
        assert all("2099" in str(item.formula2) for item in date_validations)


def test_as_of_date_duplicate_and_long_date_after_com_recalculation(
    generated_path, tmp_path
):
    path = tmp_path / "as-of-duplicate-long-date.xlsx"
    shutil.copy2(generated_path, path)
    workbook = openpyxl.load_workbook(path)
    instruction_ws = workbook["使用说明"]
    plan_ws = workbook["计划参数"]
    event_ws = workbook["激励事件明细"]
    summary_ws = workbook["年度计税汇总"]
    ledger_ws = workbook["分期缴税台账"]
    single_ws = workbook["单人测算"]
    event_headers = header_map(event_ws)
    summary_headers = header_map(summary_ws)
    ledger_headers = header_map(ledger_ws)

    instruction_ws["B21"] = date(2026, 6, 30)
    for row, employee_id, name in (
        (3, "ASOF001", "截至日测试"),
        (4, "DUP001", "重复事件测试"),
        (5, "LONG001", "长期日期测试"),
    ):
        plan_ws.cell(row, 9, employee_id)
        plan_ws.cell(row, 10, name)
    normal_event = {
        "事件编号": "ASOF-EVT-001",
        "员工编号": "ASOF001",
        "姓名": "截至日测试",
        "激励计划": "2023年股票期权与限制性股票激励计划",
        "权益类型": "股票期权",
        "事件日期": date(2026, 6, 1),
        "本次行权或解禁数量": 50000,
        "行权日或解禁日收盘价": 17.8,
        "备案状态": "已备案",
    }
    for header, value in normal_event.items():
        event_ws.cell(4, event_headers[header], value)

    for row, event_date in ((5, date(2026, 5, 1)), (6, date(2026, 5, 2))):
        duplicate_event = {
            "事件编号": "DUP-EVT-001",
            "员工编号": "DUP001",
            "姓名": "重复事件测试",
            "激励计划": "2023年股票期权与限制性股票激励计划",
            "权益类型": "其他股权激励",
            "事件日期": event_date,
            "其他需合并股权激励所得": 20000,
            "备案状态": "已备案",
        }
        for header, value in duplicate_event.items():
            event_ws.cell(row, event_headers[header], value)

    long_event = {
        "事件编号": "LONG-EVT-001",
        "员工编号": "LONG001",
        "姓名": "长期日期测试",
        "激励计划": "2023年股票期权与限制性股票激励计划",
        "权益类型": "其他股权激励",
        "事件日期": date(2096, 12, 31),
        "其他需合并股权激励所得": 10000,
        "备案状态": "已备案",
    }
    for header, value in long_event.items():
        event_ws.cell(7, event_headers[header], value)

    for row, payment_date in ((2, None), (3, date(2026, 7, 1)), (4, date(2026, 6, 30))):
        ledger_ws.cell(row, ledger_headers["员工编号"], "ASOF001")
        ledger_ws.cell(row, ledger_headers["姓名"], "截至日测试")
        ledger_ws.cell(row, ledger_headers["纳税年度"], 2026)
        ledger_ws.cell(row, ledger_headers["事件编号"], "ASOF-EVT-001")
        ledger_ws.cell(row, ledger_headers["缴税日期"], payment_date)
        ledger_ws.cell(row, ledger_headers["实际缴税金额"], 1000)
    ledger_ws.cell(5, ledger_headers["员工编号"], "DUP001")
    ledger_ws.cell(5, ledger_headers["事件编号"], "DUP-EVT-001")
    ledger_ws.cell(5, ledger_headers["缴税日期"], date(2026, 6, 1))
    ledger_ws.cell(5, ledger_headers["实际缴税金额"], 1000)
    ledger_ws.cell(6, ledger_headers["事件编号"], "ASOF-EVT-001")
    ledger_ws.cell(6, ledger_headers["缴税日期"], date(2026, 7, 15))
    ledger_ws.cell(6, ledger_headers["计划缴税金额"], 2000)

    summary_ws["A3"] = "ASOF001"
    summary_ws["B3"] = "截至日测试"
    summary_ws["C3"] = 2026
    summary_ws["A4"] = "DUP001"
    summary_ws["B4"] = "重复事件测试"
    summary_ws["C4"] = 2026
    single_ws["B3"] = "ASOF001"
    single_ws["B4"] = 2026
    workbook.save(path)

    result = excel_com_open(path, save=True)
    assert result.returncode == 0, result.stdout + result.stderr
    calculated = openpyxl.load_workbook(path, data_only=True)
    events = calculated["激励事件明细"]
    summary = calculated["年度计税汇总"]
    ledger = calculated["分期缴税台账"]
    single = calculated["单人测算"]

    effective_col = ledger_headers["截至日有效实缴"]
    assert ledger.cell(2, effective_col).value == 0
    assert ledger.cell(3, effective_col).value == 0
    assert ledger.cell(4, effective_col).value == 1000
    assert ledger.cell(6, effective_col).value == 0
    assert ledger.cell(2, ledger_headers["累计已缴"]).value == 1000
    assert "缴税日期为空" in ledger.cell(2, ledger_headers["校验"]).value
    assert "晚于截至日期" in ledger.cell(3, ledger_headers["校验"]).value
    assert ledger.cell(4, ledger_headers["校验"]).value == "正常"
    assert events.cell(4, event_headers["批次有效已缴"]).value == 1000
    assert summary.cell(3, summary_headers["已缴金额"]).value == 1000
    assert single["B8"].value == 1000

    for row in (5, 6):
        assert events.cell(row, event_headers["有效事件"]).value == "无效"
        assert events.cell(row, event_headers["原始本次所得"]).value == 0
        assert events.cell(row, event_headers["本次应纳税所得额"]).value == 0
        assert events.cell(row, event_headers["本次新增税额"]).value in (None, 0)
        assert "事件编号重复" in events.cell(row, event_headers["校验"]).value
    assert summary.cell(4, summary_headers["年度股权激励应纳税所得额"]).value == 0
    assert summary.cell(4, summary_headers["年度应纳税额"]).value == 0
    assert ledger.cell(5, ledger_headers["解析事件编号"]).value in (None, "")
    assert ledger.cell(5, effective_col).value == 0
    assert "有效事件编号或税额批次" in ledger.cell(5, ledger_headers["校验"]).value

    deadline = events.cell(7, event_headers["事件级36个月截止日"]).value
    assert deadline.date() == date(2099, 12, 31)
    assert events.cell(7, event_headers["校验"]).value == "正常"


def test_other_equity_income_requires_only_other_income_fields(workbook):
    ws = workbook["激励事件明细"]
    headers = header_map(ws)
    raw_formula = ws.cell(4, headers["原始本次所得"]).value
    validation_formula = ws.cell(4, headers["校验"]).value
    assert 'F4="其他股权激励"' in raw_formula
    assert "O4" in raw_formula
    assert 'F4<>"其他股权激励"' in validation_formula
    batches = event_tax_batches(
        [
            {
                "event_id": "OTHER-001",
                "employee_id": "E001",
                "event_date": date(2026, 6, 1),
                "income": Decimal("50000"),
            }
        ]
    )
    assert batches[0]["incremental_tax"] == Decimal("2480.00")


def test_event_prices_reference_plan_parameters_without_hardcoding(workbook):
    ws = workbook["激励事件明细"]
    headers = header_map(ws)
    formulas = "\n".join(
        ws.cell(4, headers[name]).value
        for name in (
            "适用行权或授予价格",
            "限制性股票登记日收盘价",
        )
    )
    assert formulas.count("'计划参数'!") >= 2
    for literal in ("14.56", "14.32", "14.18", "14.04", "13.06", "8.68"):
        assert literal not in formulas


def test_negative_raw_income_is_zero_and_warned(workbook):
    ws = workbook["激励事件明细"]
    headers = header_map(ws)
    raw_formula = ws.cell(4, headers["原始本次所得"]).value
    taxable_formula = ws.cell(4, headers["本次应纳税所得额"]).value
    validation_formula = ws.cell(4, headers["校验"]).value
    assert "MAX(0" not in raw_formula
    assert "MAX(0" in taxable_formula
    assert "原始所得为负" in validation_formula


def test_annual_summary_uses_non_array_helper_key_lookup(workbook):
    ws = workbook["年度计税汇总"]
    headers = header_map(ws)
    formula = ws.cell(2, headers["最早到期事件编号"]).value
    paid_formula = ws.cell(2, headers["已缴金额"]).value
    assert "MATCH(1,(" not in formula
    assert "INDEX" in formula and "MATCH" in formula
    assert "年度到期键" in header_map(workbook["激励事件明细"])
    assert "年度未缴到期键" in header_map(workbook["激励事件明细"])
    assert "O2" not in formula
    assert "'分期缴税台账'!$Q$2:$Q$501" in paid_formula
    assert '"有效"' in paid_formula


def test_demo_annual_summary_has_recalculated_earliest_unpaid_event(recalculated_path):
    workbook = openpyxl.load_workbook(recalculated_path, data_only=True)
    ws = workbook["年度计税汇总"]
    assert ws["N2"].value == "DEMO-OPT-001"
    assert ws["O2"].value.date() == date(2028, 1, 9)
    assert ws["P2"].value == "未缴清"


def test_multiple_events_select_earliest_unpaid_deadline(generated_path, tmp_path):
    path = tmp_path / "multiple-events.xlsx"
    shutil.copy2(generated_path, path)
    workbook = openpyxl.load_workbook(path)
    event_ws = workbook["激励事件明细"]
    event_headers = header_map(event_ws)
    plan_ws = workbook["计划参数"]
    plan_ws["I3"] = "MULTI001"
    plan_ws["J3"] = "多事件员工"
    events = [
        ("MULTI-LATE", date(2025, 8, 31), 50000, 18.0),
        ("MULTI-EARLY", date(2025, 2, 28), 50000, 18.0),
    ]
    for row, (event_id, event_date, quantity, market_price) in enumerate(events, start=4):
        values = {
            "事件编号": event_id,
            "员工编号": "MULTI001",
            "姓名": "多事件员工",
            "激励计划": "2023年股票期权与限制性股票激励计划",
            "权益类型": "股票期权",
            "事件日期": event_date,
            "本次行权或解禁数量": quantity,
            "行权日或解禁日收盘价": market_price,
            "备案状态": "已备案",
        }
        for header, value in values.items():
            event_ws.cell(row, event_headers[header], value)
    summary_ws = workbook["年度计税汇总"]
    summary_ws["A3"] = "MULTI001"
    summary_ws["C3"] = 2025
    ledger_ws = workbook["分期缴税台账"]
    ledger_headers = header_map(ledger_ws)
    ledger_ws.cell(2, ledger_headers["事件编号"], "MULTI-EARLY")
    ledger_ws.cell(2, ledger_headers["缴税日期"], date(2025, 3, 15))
    ledger_ws.cell(2, ledger_headers["实际缴税金额"], 19880)
    workbook.save(path)

    result = excel_com_open(path, save=True)
    assert result.returncode == 0, result.stdout + result.stderr
    calculated = openpyxl.load_workbook(path, data_only=True)
    summary = calculated["年度计税汇总"]
    assert summary["N3"].value == "MULTI-LATE"
    assert summary["O3"].value.date() == date(2028, 8, 31)
    assert summary["P3"].value == "未缴清"


def test_single_person_sheet_only_references_source_sheets(workbook):
    ws = workbook["单人测算"]
    formulas = "\n".join(cell.value for cell in formula_cells(ws))
    assert "激励事件明细" in formulas
    assert "年度计税汇总" in formulas
    assert "分期缴税台账" in formulas
    assert "VLOOKUP" not in formulas or "税率及规则" not in formulas
    values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
    for heading in ("事件明细", "分期记录", "批次余额", "到期/逾期提示"):
        assert heading in values
    assert ws["B3"].data_type != "f"
    assert "'分期缴税台账'!$Q$2:$Q$501" in ws["B8"].value
    validation = next(
        item for item in ws.data_validations.dataValidation if "B3" in str(item.sqref)
    )
    assert validation.formula1 == "=员工主数据编号"
    for cell in ("B12", "D12", "F12"):
        assert ws[cell].value == 1
        assert not ws[cell].protection.locked
    for cell in ("B13", "D13", "F13"):
        assert "ROUNDUP" in ws[cell].value
        assert "/25" in ws[cell].value
    assert "25" in ws["A25"].value or "25" in ws["A25"].value.replace("$", "")
    formulas = "\n".join(cell.value for cell in formula_cells(ws))
    assert "解析事件编号" not in formulas
    assert "'分期缴税台账'!$W$2:$W$501" in formulas
    assert "越界" in formulas


def test_validations_styles_protection_and_print_settings(workbook):
    event_ws = workbook["激励事件明细"]
    ledger_ws = workbook["分期缴税台账"]
    summary_ws = workbook["年度计税汇总"]
    event_validations = event_ws.data_validations.dataValidation
    assert any("C2:C501" in str(validation.sqref) for validation in event_validations)
    assert len(ledger_ws.data_validations.dataValidation) >= 3

    for ws in (event_ws, ledger_ws, summary_ws):
        assert ws.freeze_panes
        assert ws.auto_filter.ref or ws.tables
        assert ws.print_area
        assert ws.sheet_properties.pageSetUpPr.fitToPage
        assert ws.protection.sheet

    headers = header_map(event_ws)
    input_cell = event_ws.cell(4, headers["员工编号"])
    formula_cell = event_ws.cell(4, headers["纳税年度"])
    assert input_cell.fill.fgColor.rgb.endswith("DDEBF7")
    assert not input_cell.protection.locked
    assert formula_cell.fill.fgColor.rgb.endswith("E2F0D9")
    assert formula_cell.protection.locked
    assert event_ws.conditional_formatting


def test_no_formula_contains_excel_error_tokens(workbook):
    for ws in workbook.worksheets:
        for cell in formula_cells(ws):
            assert not any(token in cell.value for token in ERROR_TOKENS), (
                ws.title,
                cell.coordinate,
                cell.value,
            )


def test_xml_has_no_duplicate_sheet_autofilter_or_aggregate_formula(generated_path):
    with zipfile.ZipFile(generated_path) as archive:
        for name in archive.namelist():
            if not name.startswith("xl/worksheets/sheet") or not name.endswith(".xml"):
                continue
            root = ElementTree.fromstring(archive.read(name))
            has_table_parts = root.find("m:tableParts", XML_NS) is not None
            if has_table_parts:
                assert root.find("m:autoFilter", XML_NS) is None
            for formula in root.findall(".//m:f", XML_NS):
                assert "AGGREGATE(" not in (formula.text or "")
                assert "MINIFS(" not in (formula.text or "")
