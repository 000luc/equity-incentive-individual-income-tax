from __future__ import annotations

import os
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = Path(
    os.environ.get(
        "WORKBOOK_OUTPUT",
        ROOT / "广电计量股权激励个人所得税计算及分期台账-2026版.xlsx",
    )
)
EVENT_LAST_ROW = 501
LEDGER_LAST_ROW = 501
SUMMARY_LAST_ROW = 201

NAVY = "1F4E78"
WHITE = "FFFFFF"
INPUT_BLUE = "DDEBF7"
PARAM_GRAY = "D9E1F2"
FORMULA_GREEN = "E2F0D9"
ALERT_RED = "F4CCCC"
NOTE_YELLOW = "FFF2CC"
THIN_GRAY = Side(style="thin", color="B7B7B7")
MONEY_FORMAT = '#,##0.00;[Red]-#,##0.00'
DATE_FORMAT = "yyyy-mm-dd"


def style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN_GRAY)
    ws.row_dimensions[row].height = 32


def set_column_widths(ws, widths: dict[int, float]) -> None:
    for column, width in widths.items():
        ws.column_dimensions[get_column_letter(column)].width = width


def add_table(ws, name: str, last_row: int, last_col: int) -> None:
    table = Table(displayName=name, ref=f"A1:{get_column_letter(last_col)}{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def set_print(ws, last_row: int, last_col: int, landscape: bool = True) -> None:
    ws.print_area = f"A1:{get_column_letter(last_col)}{last_row}"
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:1"
    ws.sheet_properties.outlinePr.summaryBelow = True


def fill_cells(ws, columns: set[int], start_row: int, end_row: int, color: str, locked: bool) -> None:
    fill = PatternFill("solid", fgColor=color)
    for row in range(start_row, end_row + 1):
        for column in columns:
            cell = ws.cell(row, column)
            cell.fill = fill
            cell.protection = Protection(locked=locked)


def protect(ws) -> None:
    ws.protection.sheet = True
    ws.protection.autoFilter = False
    ws.protection.sort = False
    ws.protection.selectLockedCells = True
    ws.protection.selectUnlockedCells = True


def add_list_validation(ws, formula: str, ranges: list[str]) -> None:
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    validation.error = "请选择下拉列表中的有效值。"
    validation.errorTitle = "输入无效"
    validation.prompt = "请从下拉列表选择。"
    validation.promptTitle = "可选值"
    validation.showErrorMessage = True
    validation.showInputMessage = True
    ws.add_data_validation(validation)
    for cell_range in ranges:
        validation.add(cell_range)


def add_decimal_validation(
    ws,
    ranges: list[str],
    operator: str = "greaterThanOrEqual",
    value: str = "0",
    value2: str | None = None,
) -> None:
    validation = DataValidation(
        type="decimal",
        operator=operator,
        formula1=value,
        formula2=value2,
        allow_blank=True,
        showErrorMessage=True,
    )
    validation.error = "请输入符合条件的非负数字。"
    validation.errorTitle = "数值无效"
    ws.add_data_validation(validation)
    for cell_range in ranges:
        validation.add(cell_range)


def add_date_validation(ws, ranges: list[str]) -> None:
    validation = DataValidation(
        type="date",
        operator="between",
        formula1="DATE(2024,1,1)",
        formula2="DATE(2035,12,31)",
        allow_blank=True,
        showErrorMessage=True,
    )
    validation.error = "请输入2024-01-01至2035-12-31之间的日期。"
    validation.errorTitle = "日期无效"
    ws.add_data_validation(validation)
    for cell_range in ranges:
        validation.add(cell_range)


def create_instructions(wb: Workbook) -> None:
    ws = wb.create_sheet("使用说明")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    ws["A1"] = "广电计量股权激励个人所得税计算及分期台账（2026版）"
    ws["A1"].font = Font(size=16, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center")
    rows = [
        ("一、录入顺序", "先维护计划参数，再逐笔录入激励事件，最后在分期缴税台账登记计划及实际缴税。"),
        ("二、颜色", "蓝色为人工输入；灰色为公告参数；绿色为公式；红色为异常；黄色为说明。"),
        ("三、年度计税", "同一员工同一纳税年度的股权激励所得合并计税，按事件日期及事件编号稳定排序逐笔确认新增税额。"),
        ("四、36个月", "每次事件新增税额独立形成批次，从该事件行权日或解禁日起单独计算36个月。"),
        ("五、限制性股票", "个人获授总数和个人实际出资必须按员工口径录入，公司802万股绝不能作为个人分母。"),
        ("六、员工主数据", "先在计划参数表I:J列维护员工编号和姓名；事件明细员工编号可下拉选择，也允许新增后补录主数据。"),
        ("七、分期缴税", "缴税日期、计划金额和实际金额均可自定义；事件编号或税额批次任填其一，双填时必须指向同一事件。"),
        ("八、其他股权激励", "选择“其他股权激励”时，只需填写其他需合并股权激励所得，不要求数量和市场价。"),
        ("九、单人测算", "该页只引用事件明细、年度汇总和分期台账；事件、缴税、批次余额各25行分页，最多20页。"),
        ("十、范围", "不计算股息红利个人所得税；演示数据仅用于验证公式，不得直接用于申报。"),
        ("十一、保护", "公式及公告参数已启用无密码保护；需要维护时可在Excel“审阅-撤销工作表保护”后操作。"),
        ("十二、申报", "36个月缴税以完成主管税务机关备案为前提，最终申报口径以备案及主管税务机关要求为准。"),
    ]
    for index, (title, body) in enumerate(rows, start=3):
        ws.cell(index, 1, title).font = Font(bold=True, color=NAVY)
        ws.cell(index, 2, body)
        ws.merge_cells(start_row=index, start_column=2, end_row=index, end_column=8)
        ws.cell(index, 2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[index].height = 34
    ws["A17"] = "演示员工"
    ws["B17"] = "DEMO001，两条事件均标注“仅演示”。"
    ws["A19"] = "图例"
    for col, (label, color) in enumerate(
        [("人工输入", INPUT_BLUE), ("公告参数", PARAM_GRAY), ("公式结果", FORMULA_GREEN), ("异常提示", ALERT_RED)],
        start=2,
    ):
            ws.cell(19, col, label)
            ws.cell(19, col).fill = PatternFill("solid", fgColor=color)
    set_column_widths(ws, {1: 18, 2: 24, 3: 18, 4: 18, 5: 18, 6: 18, 7: 18, 8: 18})
    ws.print_area = "A1:H20"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    protect(ws)


def create_plan_parameters(wb: Workbook) -> None:
    ws = wb.create_sheet("计划参数")
    headers = ["参数类别", "参数名称", "参数值", "生效日", "失效日", "口径/说明", "来源公告"]
    ws.append(headers)
    rows = [
        ("基本", "激励计划", "2023年股票期权与限制性股票激励计划", None, None, "2023计划", "2024-035"),
        ("基本", "股票代码", "002967", None, None, "广电计量", "公司公告"),
        ("实际授予", "股票期权实际授予数量", 8625000, date(2024, 7, 1), None, "623人；授予决定口径", "2024-043"),
        ("完成登记", "股票期权登记数量", 8020000, date(2024, 7, 24), None, "557人；完成登记口径", "2024-045"),
        ("实际授予", "限制性股票实际授予数量", 8625000, date(2024, 7, 1), None, "623人；授予决定口径", "2024-043"),
        ("完成登记", "限制性股票登记数量", 8020000, date(2024, 8, 27), None, "557人；完成登记口径", "2024-054"),
        ("限制性股票", "登记日", date(2024, 8, 27), date(2024, 8, 27), None, "登记完成日及上市日", "2024-054及后续公告"),
        ("限制性股票", "登记日收盘价", 13.06, date(2024, 8, 27), None, "深交所2024-08-27今收", "深交所历史行情"),
        ("限制性股票", "授予价格", 8.68, date(2024, 7, 1), None, "仅用于核对个人实际出资", "2024-043"),
        ("个人口径", "限制性股票个人分母", "员工实际登记取得总股数", None, None, "802万股不得作为个人分母", "国税函〔2009〕461号"),
        ("行权价格", "股票期权行权价格", 14.56, date(2024, 7, 1), date(2025, 1, 9), "按事件日自动选择", "2024-043/045"),
        ("行权价格", "股票期权行权价格", 14.32, date(2025, 1, 10), date(2025, 4, 28), "按事件日自动选择", "2025-002"),
        ("行权价格", "股票期权行权价格", 14.18, date(2025, 4, 29), date(2025, 9, 25), "按事件日自动选择", "2025-031"),
        ("行权价格", "股票期权行权价格", 14.04, date(2025, 9, 26), None, "按本任务要求内置至该区间", "2025-060"),
    ]
    for row in rows:
        ws.append(row)
    style_header(ws)
    fill_cells(ws, set(range(1, 8)), 2, ws.max_row, PARAM_GRAY, True)
    ws["I1"] = "员工编号"
    ws["J1"] = "姓名"
    for cell in (ws["I1"], ws["J1"]):
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center")
    ws["I2"] = "DEMO001"
    ws["J2"] = "演示员工"
    fill_cells(ws, {9, 10}, 2, EVENT_LAST_ROW, INPUT_BLUE, False)
    ws["I502"] = "说明"
    ws["J502"] = "员工编号和姓名主数据输入区；可继续新增，事件明细下拉引用本区域。"
    ws["I502"].fill = PatternFill("solid", fgColor=NOTE_YELLOW)
    ws["J502"].fill = PatternFill("solid", fgColor=NOTE_YELLOW)
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 4).number_format = DATE_FORMAT
        ws.cell(row, 5).number_format = DATE_FORMAT
        if isinstance(ws.cell(row, 3).value, float):
            ws.cell(row, 3).number_format = "0.00"
    add_table(ws, "tblPlanParameters", len(rows) + 1, len(headers))
    set_print(ws, 35, 10)
    set_column_widths(ws, {1: 16, 2: 28, 3: 28, 4: 14, 5: 14, 6: 36, 7: 24, 9: 18, 10: 18})
    protect(ws)


def event_headers() -> list[str]:
    return [
        "事件编号",
        "税额批次",
        "员工编号",
        "姓名",
        "激励计划",
        "权益类型",
        "事件日期",
        "纳税年度",
        "本次行权或解禁数量",
        "行权日或解禁日收盘价",
        "适用行权或授予价格",
        "限制性股票登记日收盘价",
        "员工获授限制性股票总数",
        "员工实际出资总额",
        "其他需合并股权激励所得",
        "离职日期",
        "备案状态",
        "备案日期",
        "原始本次所得",
        "本次应纳税所得额",
        "此前年度累计所得",
        "本次事件后年度累计所得",
        "税率",
        "速算扣除数",
        "累计税额",
        "此前累计已确认税额",
        "本次新增税额",
        "本事件纳税义务发生日",
        "事件级36个月截止日",
        "校验",
        "备注",
        "员工首次序号",
        "员工事件序号",
        "员工事件键",
        "年度到期键",
        "员工到期键",
    ]


def create_event_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("激励事件明细")
    headers = event_headers()
    ws.append(headers)
    for _ in range(2, EVENT_LAST_ROW + 1):
        ws.append([None] * len(headers))

    demo = [
        ("DEMO-OPT-001", "DEMO001", "演示员工", "股票期权", date(2025, 1, 9), 50000, 17.8, None, None, 0, "已备案", date(2025, 1, 8), "仅演示：5万份期权"),
        ("DEMO-RS-001", "DEMO001", "演示员工", "限制性股票", date(2025, 6, 30), 50000, 17.8, 50000, 434000, 0, "已备案", date(2024, 9, 10), "仅演示：5万股限制性股票"),
    ]
    for row, values in enumerate(demo, start=2):
        event_id, employee_id, name, equity_type, event_date, qty, market, total, paid, other, filing, filing_date, note = values
        inputs = {
            1: event_id,
            3: employee_id,
            4: name,
            5: "2023年股票期权与限制性股票激励计划",
            6: equity_type,
            7: event_date,
            9: qty,
            10: market,
            13: total,
            14: paid,
            15: other,
            17: filing,
            18: filing_date,
            31: note,
        }
        for column, value in inputs.items():
            ws.cell(row, column, value)

    for row in range(2, EVENT_LAST_ROW + 1):
        ws.cell(row, 2, f'=IF(A{row}="","",A{row}&"-TAX")')
        ws.cell(row, 8, f'=IF(G{row}="","",YEAR(G{row}))')
        ws.cell(
            row,
            11,
            f'=IF(OR(E{row}="",F{row}="",G{row}=""),"",IF(E{row}<>\'计划参数\'!$C$2,"",'
            f'IF(F{row}="股票期权",IFERROR(LOOKUP(G{row},\'计划参数\'!$D$12:$D$15,\'计划参数\'!$C$12:$C$15),""),'
            f'IF(F{row}="限制性股票",\'计划参数\'!$C$10,""))))',
        )
        ws.cell(
            row,
            12,
            f'=IF(AND(E{row}=\'计划参数\'!$C$2,F{row}="限制性股票",G{row}>=\'计划参数\'!$D$9),'
            f'\'计划参数\'!$C$9,"")',
        )
        ws.cell(
            row,
            19,
            f'=IF(OR(A{row}="",C{row}="",F{row}="",G{row}=""),"",'
            f'IF(F{row}="股票期权",IF(OR(I{row}="",J{row}="",K{row}=""),"",(J{row}-K{row})*I{row}),'
            f'IF(F{row}="限制性股票",IF(OR(L{row}="",M{row}="",M{row}<=0,N{row}=""),"",'
            f'((L{row}+J{row})/2)*I{row}-N{row}*I{row}/M{row}),'
            f'IF(F{row}="其他股权激励",IF(O{row}="","",O{row}),""))))',
        )
        ws.cell(
            row,
            20,
            f'=IF(S{row}="","",MAX(0,S{row}))',
        )
        ws.cell(
            row,
            21,
            f'=IF(T{row}="","",SUMIFS($T$2:$T$501,$C$2:$C$501,C{row},$H$2:$H$501,H{row},'
            f'$G$2:$G$501,"<"&G{row})+SUMIFS($T$2:$T$501,$C$2:$C$501,C{row},'
            f'$H$2:$H$501,H{row},$G$2:$G$501,G{row},$A$2:$A$501,"<"&A{row}))',
        )
        ws.cell(row, 22, f'=IF(T{row}="","",U{row}+T{row})')
        ws.cell(row, 23, f'=IF(V{row}="","",VLOOKUP(V{row},\'税率及规则\'!$A$2:$C$8,2,TRUE))')
        ws.cell(row, 24, f'=IF(V{row}="","",VLOOKUP(V{row},\'税率及规则\'!$A$2:$C$8,3,TRUE))')
        ws.cell(row, 25, f'=IF(V{row}="","",MAX(0,V{row}*W{row}-X{row}))')
        ws.cell(
            row,
            26,
            f'=IF(U{row}="","",MAX(0,U{row}*VLOOKUP(U{row},\'税率及规则\'!$A$2:$C$8,2,TRUE)-'
            f'VLOOKUP(U{row},\'税率及规则\'!$A$2:$C$8,3,TRUE)))',
        )
        ws.cell(row, 27, f'=IF(Y{row}="","",MAX(0,Y{row}-Z{row}))')
        ws.cell(row, 28, f'=IF(G{row}="","",G{row})')
        ws.cell(row, 29, f'=IF(AB{row}="","",EDATE(AB{row},36))')
        ws.cell(
            row,
            30,
            f'=IF(A{row}="","",IF(COUNTIF($A$2:$A$501,A{row})>1,"异常：事件编号重复",'
            f'IF(OR(C{row}="",D{row}="",E{row}="",F{row}="",G{row}=""),"异常：必填项缺失",'
            f'IF(AND(F{row}="其他股权激励",O{row}=""),"异常：其他股权激励所得缺失",'
            f'IF(AND(F{row}<>"其他股权激励",OR(I{row}="",J{row}="")),"异常：数量或市场价缺失",'
            f'IF(AND(F{row}<>"其他股权激励",OR(I{row}<=0,J{row}<0)),"异常：数量或价格无效",'
            f'IF(AND(F{row}="限制性股票",OR(M{row}<=0,N{row}<0,I{row}>M{row})),"异常：个人获授或出资数据无效",'
            f'IF(AND(IFERROR(INDEX(\'计划参数\'!$J$2:$J$501,MATCH(C{row},\'计划参数\'!$I$2:$I$501,0)),"")<>"",'
            f'D{row}<>IFERROR(INDEX(\'计划参数\'!$J$2:$J$501,MATCH(C{row},\'计划参数\'!$I$2:$I$501,0)),"")),"异常：姓名与员工主数据不一致",'
            f'IF(S{row}<0,"异常：原始所得为负，已按0计税",'
            f'IF(OR(AND(F{row}="股票期权",K{row}=""),AND(F{row}="限制性股票",OR(K{row}="",L{row}=""))),"异常：计划参数缺失",'
            f'IF(COUNTIF(\'计划参数\'!$I$2:$I$501,C{row})=0,"提示：员工编号未在主数据",'
            f'IF(Q{row}<>"已备案","提示：未确认备案","正常"))))))))))))',
        )
        ws.cell(row, 32, f'=IF(C{row}="","",IF(COUNTIF($C$2:C{row},C{row})=1,MAX($AF$1:AF{row - 1})+1,""))')
        ws.cell(row, 33, f'=IF(C{row}="","",COUNTIF($C$2:C{row},C{row}))')
        ws.cell(row, 34, f'=IF(C{row}="","",C{row}&"|"&AG{row})')
        ws.cell(
            row,
            35,
            f'=IF(OR(A{row}="",AA{row}<=0),"",C{row}&"|"&H{row}&"|"&('
            f'COUNTIFS($C$2:$C$501,C{row},$H$2:$H$501,H{row},$AA$2:$AA$501,">0",$AC$2:$AC$501,"<"&AC{row})+'
            f'COUNTIFS($C$2:$C$501,C{row},$H$2:$H$501,H{row},$AA$2:$AA$501,">0",$AC$2:$AC$501,AC{row},$A$2:$A$501,"<="&A{row})))',
        )
        ws.cell(
            row,
            36,
            f'=IF(OR(A{row}="",AA{row}<=0),"",C{row}&"|"&('
            f'COUNTIFS($C$2:$C$501,C{row},$AA$2:$AA$501,">0",$AC$2:$AC$501,"<"&AC{row})+'
            f'COUNTIFS($C$2:$C$501,C{row},$AA$2:$AA$501,">0",$AC$2:$AC$501,AC{row},$A$2:$A$501,"<="&A{row})))',
        )

    style_header(ws)
    input_columns = {1, 3, 4, 5, 6, 7, 9, 10, 13, 14, 15, 16, 17, 18, 31}
    formula_columns = set(range(1, len(headers) + 1)) - input_columns
    fill_cells(ws, input_columns, 2, EVENT_LAST_ROW, INPUT_BLUE, False)
    fill_cells(ws, formula_columns, 2, EVENT_LAST_ROW, FORMULA_GREEN, True)
    for row in range(2, EVENT_LAST_ROW + 1):
        for column in (7, 16, 18, 28, 29):
            ws.cell(row, column).number_format = DATE_FORMAT
        for column in (10, 11, 12, 14, 15, 19, 20, 21, 22, 24, 25, 26, 27):
            ws.cell(row, column).number_format = MONEY_FORMAT
        ws.cell(row, 23).number_format = "0%"
    add_list_validation(ws, '"2023年股票期权与限制性股票激励计划"', [f"E2:E{EVENT_LAST_ROW}"])
    add_list_validation(ws, '"股票期权,限制性股票,其他股权激励"', [f"F2:F{EVENT_LAST_ROW}"])
    add_list_validation(ws, '"已备案,未备案,待确认"', [f"Q2:Q{EVENT_LAST_ROW}"])
    employee_validation = DataValidation(
        type="list",
        formula1="=员工主数据编号",
        allow_blank=True,
        errorStyle="information",
        showErrorMessage=True,
    )
    employee_validation.errorTitle = "员工编号未在主数据"
    employee_validation.error = "可继续录入，但请同步补充计划参数表员工主数据。"
    ws.add_data_validation(employee_validation)
    employee_validation.add(f"C2:C{EVENT_LAST_ROW}")
    add_date_validation(ws, [f"G2:G{EVENT_LAST_ROW}", f"P2:P{EVENT_LAST_ROW}", f"R2:R{EVENT_LAST_ROW}"])
    add_decimal_validation(ws, [f"I2:I{EVENT_LAST_ROW}"], "greaterThan", "0")
    add_decimal_validation(ws, [f"J2:J{EVENT_LAST_ROW}", f"M2:O{EVENT_LAST_ROW}"])
    ws.conditional_formatting.add(
        f"AD2:AD{EVENT_LAST_ROW}",
        FormulaRule(formula=['LEFT(AD2,2)="异常"'], fill=PatternFill("solid", fgColor=ALERT_RED)),
    )
    add_table(ws, "tblIncentiveEvents", EVENT_LAST_ROW, len(headers))
    set_print(ws, EVENT_LAST_ROW, len(headers))
    set_column_widths(
        ws,
        {
            1: 18, 2: 22, 3: 14, 4: 12, 5: 30, 6: 16, 7: 13, 8: 10, 9: 18, 10: 20,
            11: 20, 12: 22, 13: 22, 14: 20, 15: 22, 16: 13, 17: 12, 18: 13, 19: 20,
            20: 20, 21: 22, 22: 10, 23: 14, 24: 16, 25: 20, 26: 16, 27: 18, 28: 22,
            29: 28, 30: 28,
        },
    )
    for column in ("AF", "AG", "AH", "AI", "AJ"):
        ws.column_dimensions[column].hidden = True
    protect(ws)


def create_tax_rules(wb: Workbook) -> None:
    ws = wb.create_sheet("税率及规则")
    ws.append(["所得下限", "税率", "速算扣除数", "所得上限", "规则说明"])
    brackets = [
        (0, 0.03, 0, 36000),
        (36000.01, 0.10, 2520, 144000),
        (144000.01, 0.20, 16920, 300000),
        (300000.01, 0.25, 31920, 420000),
        (420000.01, 0.30, 52920, 660000),
        (660000.01, 0.35, 85920, 960000),
        (960000.01, 0.45, 181920, "以上"),
    ]
    for lower, rate, deduction, upper in brackets:
        ws.append([lower, rate, deduction, upper, "同一员工同一年度股权激励所得合并后单独计税"])
    ws["A10"] = "股票期权所得"
    ws["B10"] = "（行权日收盘价－事件日有效行权价）×本次行权数量"
    ws["A11"] = "限制性股票所得"
    ws["B11"] = "（登记日收盘价＋解禁日收盘价）÷2×本批解禁数量－个人实际出资×本批数量÷个人获授总数"
    ws["A12"] = "36个月"
    ws["B12"] = "每次事件新增税额独立起算，不按年度最早事件统一截止"
    ws["A13"] = "月末规则"
    ws["B13"] = "使用EDATE按公历加36个月"
    ws["A14"] = "离职"
    ws["B14"] = "在36个月期间内离职的，应在离职前缴清全部税款"
    style_header(ws)
    fill_cells(ws, {1, 2, 3, 4, 5}, 2, ws.max_row, PARAM_GRAY, True)
    for row in range(2, 9):
        ws.cell(row, 1).number_format = MONEY_FORMAT
        ws.cell(row, 2).number_format = "0%"
        ws.cell(row, 3).number_format = MONEY_FORMAT
    add_table(ws, "tblTaxRates", 8, 5)
    set_print(ws, ws.max_row, 5)
    set_column_widths(ws, {1: 18, 2: 12, 3: 18, 4: 18, 5: 55})
    protect(ws)


def create_summary(wb: Workbook) -> None:
    ws = wb.create_sheet("年度计税汇总")
    headers = [
        "员工编号", "姓名", "纳税年度", "股票期权所得", "限制性股票所得", "其他股权激励所得",
        "年度股权激励应纳税所得额", "税率", "速算扣除数", "年度应纳税额", "已缴金额", "未缴金额",
        "事件新增税额批次数", "最早到期事件编号", "最早到期日", "状态",
    ]
    ws.append(headers)
    for _ in range(2, SUMMARY_LAST_ROW + 1):
        ws.append([None] * len(headers))
    ws["A2"] = "DEMO001"
    ws["C2"] = 2025
    for row in range(2, SUMMARY_LAST_ROW + 1):
        ws.cell(row, 2, f'=IF(A{row}="","",IFERROR(INDEX(\'激励事件明细\'!$D$2:$D$501,MATCH(A{row},\'激励事件明细\'!$C$2:$C$501,0)),""))')
        ws.cell(row, 4, f'=IF(OR(A{row}="",C{row}=""),"",SUMIFS(\'激励事件明细\'!$T$2:$T$501,\'激励事件明细\'!$C$2:$C$501,A{row},\'激励事件明细\'!$H$2:$H$501,C{row},\'激励事件明细\'!$F$2:$F$501,"股票期权"))')
        ws.cell(row, 5, f'=IF(OR(A{row}="",C{row}=""),"",SUMIFS(\'激励事件明细\'!$T$2:$T$501,\'激励事件明细\'!$C$2:$C$501,A{row},\'激励事件明细\'!$H$2:$H$501,C{row},\'激励事件明细\'!$F$2:$F$501,"限制性股票"))')
        ws.cell(row, 6, f'=IF(OR(A{row}="",C{row}=""),"",SUMIFS(\'激励事件明细\'!$T$2:$T$501,\'激励事件明细\'!$C$2:$C$501,A{row},\'激励事件明细\'!$H$2:$H$501,C{row},\'激励事件明细\'!$F$2:$F$501,"其他股权激励"))')
        ws.cell(row, 7, f'=IF(A{row}="","",SUM(D{row}:F{row}))')
        ws.cell(row, 8, f'=IF(G{row}="","",VLOOKUP(G{row},\'税率及规则\'!$A$2:$C$8,2,TRUE))')
        ws.cell(row, 9, f'=IF(G{row}="","",VLOOKUP(G{row},\'税率及规则\'!$A$2:$C$8,3,TRUE))')
        ws.cell(row, 10, f'=IF(G{row}="","",MAX(0,G{row}*H{row}-I{row}))')
        ws.cell(row, 11, f'=IF(A{row}="","",SUMIFS(\'分期缴税台账\'!$K$2:$K$501,\'分期缴税台账\'!$A$2:$A$501,A{row},\'分期缴税台账\'!$C$2:$C$501,C{row},\'分期缴税台账\'!$Q$2:$Q$501,"有效"))')
        ws.cell(row, 12, f'=IF(J{row}="","",MAX(0,J{row}-K{row}))')
        ws.cell(row, 13, f'=IF(A{row}="","",COUNTIFS(\'激励事件明细\'!$C$2:$C$501,A{row},\'激励事件明细\'!$H$2:$H$501,C{row},\'激励事件明细\'!$AA$2:$AA$501,">0"))')
        ws.cell(row, 14, f'=IF(O{row}="","",IFERROR(INDEX(\'激励事件明细\'!$A$2:$A$501,MATCH(A{row}&"|"&C{row}&"|1",\'激励事件明细\'!$AI$2:$AI$501,0)),""))')
        ws.cell(row, 15, f'=IF(N{row}="","",IFERROR(INDEX(\'激励事件明细\'!$AC$2:$AC$501,MATCH(N{row},\'激励事件明细\'!$A$2:$A$501,0)),""))')
        ws.cell(row, 16, f'=IF(A{row}="","",IF(K{row}>J{row},"异常：超额缴税",IF(L{row}=0,"已缴清",IF(AND(O{row}<TODAY(),L{row}>0),"逾期未缴","未缴清"))))')
    style_header(ws)
    fill_cells(ws, {1, 3}, 2, SUMMARY_LAST_ROW, INPUT_BLUE, False)
    fill_cells(ws, set(range(2, len(headers) + 1)) - {3}, 2, SUMMARY_LAST_ROW, FORMULA_GREEN, True)
    add_list_validation(ws, "=员工编号列表", [f"A2:A{SUMMARY_LAST_ROW}"])
    add_decimal_validation(ws, [f"C2:C{SUMMARY_LAST_ROW}"], "between", "2024", "2035")
    for row in range(2, SUMMARY_LAST_ROW + 1):
        for column in range(4, 13):
            ws.cell(row, column).number_format = MONEY_FORMAT
        ws.cell(row, 8).number_format = "0%"
        ws.cell(row, 15).number_format = DATE_FORMAT
    ws.conditional_formatting.add(
        f"P2:P{SUMMARY_LAST_ROW}",
        FormulaRule(formula=['LEFT(P2,2)="异常"'], fill=PatternFill("solid", fgColor=ALERT_RED)),
    )
    add_table(ws, "tblAnnualSummary", SUMMARY_LAST_ROW, len(headers))
    set_print(ws, SUMMARY_LAST_ROW, len(headers))
    set_column_widths(ws, {1: 14, 2: 12, 3: 10, 4: 18, 5: 20, 6: 20, 7: 24, 8: 10, 9: 14, 10: 18, 11: 16, 12: 16, 13: 18, 14: 22, 15: 14, 16: 18})
    protect(ws)


def create_ledger(wb: Workbook) -> None:
    ws = wb.create_sheet("分期缴税台账")
    headers = [
        "员工编号", "姓名", "纳税年度", "事件编号", "新增税额批次", "该事件新增应纳税额",
        "该事件纳税义务发生日", "该事件最晚缴清日", "缴税日期", "计划缴税金额", "实际缴税金额",
        "完税凭证号", "累计已缴", "剩余税额", "是否逾期", "离职日期", "关联有效", "校验", "备注",
        "解析事件编号", "解析税额批次", "员工缴税序号", "员工缴税键",
    ]
    ws.append(headers)
    for _ in range(2, LEDGER_LAST_ROW + 1):
        ws.append([None] * len(headers))
    for row in range(2, LEDGER_LAST_ROW + 1):
        ws.cell(
            row,
            20,
            f'=IF(D{row}<>"",IF(COUNTIF(\'激励事件明细\'!$A$2:$A$501,D{row})=1,D{row},""),'
            f'IF(E{row}<>"",IFERROR(INDEX(\'激励事件明细\'!$A$2:$A$501,'
            f'MATCH(E{row},\'激励事件明细\'!$B$2:$B$501,0)),""),""))',
        )
        ws.cell(
            row,
            21,
            f'=IF(T{row}="","",IFERROR(INDEX(\'激励事件明细\'!$B$2:$B$501,'
            f'MATCH(T{row},\'激励事件明细\'!$A$2:$A$501,0)),""))',
        )
        lookup = f'MATCH(T{row},\'激励事件明细\'!$A$2:$A$501,0)'
        ws.cell(row, 1, f'=IF(T{row}="","",IFERROR(INDEX(\'激励事件明细\'!$C$2:$C$501,{lookup}),""))')
        ws.cell(row, 2, f'=IF(A{row}="","",IFERROR(INDEX(\'激励事件明细\'!$D$2:$D$501,{lookup}),""))')
        ws.cell(row, 3, f'=IF(A{row}="","",IFERROR(INDEX(\'激励事件明细\'!$H$2:$H$501,{lookup}),""))')
        ws.cell(row, 6, f'=IF(A{row}="","",IFERROR(INDEX(\'激励事件明细\'!$AA$2:$AA$501,{lookup}),""))')
        ws.cell(row, 7, f'=IF(A{row}="","",IFERROR(INDEX(\'激励事件明细\'!$AB$2:$AB$501,{lookup}),""))')
        ws.cell(row, 8, f'=IF(A{row}="","",IFERROR(INDEX(\'激励事件明细\'!$AC$2:$AC$501,{lookup}),""))')
        ws.cell(
            row,
            17,
            f'=IF(AND(D{row}="",E{row}=""),"无效",IF(T{row}="","无效",'
            f'IF(AND(D{row}<>"",E{row}<>"",U{row}<>E{row}),"冲突","有效")))',
        )
        ws.cell(
            row,
            13,
            f'=IF(T{row}="","",SUMIFS($K$2:$K$501,$T$2:$T$501,T{row},$Q$2:$Q$501,"有效"))',
        )
        ws.cell(row, 14, f'=IF(F{row}="","",MAX(0,F{row}-M{row}))')
        ws.cell(row, 15, f'=IF(A{row}="","",IF(AND(I{row}<>"",I{row}>H{row}),"是",IF(AND(TODAY()>H{row},N{row}>0),"是","否")))')
        ws.cell(row, 16, f'=IF(A{row}="","",IFERROR(INDEX(\'激励事件明细\'!$P$2:$P$501,{lookup}),""))')
        ws.cell(
            row,
            18,
            f'=IF(Q{row}="冲突","异常：事件编号与税额批次冲突",'
            f'IF(Q{row}="无效","异常：必须关联有效事件编号或税额批次",'
            f'IF(AND(P{row}<>"",I{row}<>"",K{row}>0,I{row}>P{row}),"异常：离职后缴税",'
            f'IF(M{row}>F{row},"异常：本批次超额缴税",IF(AND(I{row}<>"",I{row}>H{row}),"异常：超过事件截止日",'
            f'IF(AND(P{row}<>"",TODAY()>=P{row},N{row}>0),"异常：离职前未缴清","正常"))))))',
        )
        ws.cell(row, 22, f'=IF(A{row}="","",COUNTIF($A$2:A{row},A{row}))')
        ws.cell(row, 23, f'=IF(A{row}="","",A{row}&"|"&V{row})')
    style_header(ws)
    input_columns = {4, 5, 9, 10, 11, 12, 19}
    formula_columns = set(range(1, len(headers) + 1)) - input_columns
    fill_cells(ws, input_columns, 2, LEDGER_LAST_ROW, INPUT_BLUE, False)
    fill_cells(ws, formula_columns, 2, LEDGER_LAST_ROW, FORMULA_GREEN, True)
    add_list_validation(ws, "=事件编号列表", [f"D2:D{LEDGER_LAST_ROW}"])
    add_list_validation(ws, "=税额批次列表", [f"E2:E{LEDGER_LAST_ROW}"])
    add_date_validation(ws, [f"I2:I{LEDGER_LAST_ROW}"])
    add_decimal_validation(ws, [f"J2:K{LEDGER_LAST_ROW}"])
    for row in range(2, LEDGER_LAST_ROW + 1):
        for column in (7, 8, 9, 16):
            ws.cell(row, column).number_format = DATE_FORMAT
        for column in (6, 10, 11, 13, 14):
            ws.cell(row, column).number_format = MONEY_FORMAT
    ws.conditional_formatting.add(
        f"R2:R{LEDGER_LAST_ROW}",
        FormulaRule(formula=['LEFT(R2,2)="异常"'], fill=PatternFill("solid", fgColor=ALERT_RED)),
    )
    add_table(ws, "tblInstallments", LEDGER_LAST_ROW, len(headers))
    set_print(ws, LEDGER_LAST_ROW, len(headers))
    set_column_widths(ws, {1: 14, 2: 12, 3: 10, 4: 18, 5: 22, 6: 22, 7: 22, 8: 20, 9: 14, 10: 16, 11: 16, 12: 20, 13: 16, 14: 16, 15: 12, 16: 14, 17: 28, 18: 24})
    for column in ("T", "U", "V", "W"):
        ws.column_dimensions[column].hidden = True
    protect(ws)


def create_person_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("单人测算")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "单人测算（仅引用业务明细）"
    ws["A1"].font = Font(size=15, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells("A1:G1")
    ws["A3"] = "员工编号"
    ws["B3"] = "DEMO001"
    ws["B3"].fill = PatternFill("solid", fgColor=INPUT_BLUE)
    ws["B3"].protection = Protection(locked=False)
    labels = ["姓名", "全部事件所得", "年度汇总税额", "实际已缴", "剩余税额", "最早事件截止日"]
    for row, label in enumerate(labels, start=5):
        ws.cell(row, 1, label).font = Font(bold=True)
    ws["B5"] = '=IFERROR(INDEX(\'激励事件明细\'!$D$2:$D$501,MATCH(B3,\'激励事件明细\'!$C$2:$C$501,0)),"")'
    ws["B6"] = '=SUMIFS(\'激励事件明细\'!$T$2:$T$501,\'激励事件明细\'!$C$2:$C$501,B3)'
    ws["B7"] = '=SUMIFS(\'年度计税汇总\'!$J$2:$J$201,\'年度计税汇总\'!$A$2:$A$201,B3)'
    ws["B8"] = '=SUMIFS(\'分期缴税台账\'!$K$2:$K$501,\'分期缴税台账\'!$A$2:$A$501,B3,\'分期缴税台账\'!$Q$2:$Q$501,"有效")'
    ws["B9"] = "=MAX(0,B7-B8)"
    ws["B10"] = '=IFERROR(INDEX(\'激励事件明细\'!$AC$2:$AC$501,MATCH(B3&"|1",\'激励事件明细\'!$AJ$2:$AJ$501,0)),"")'
    for row in range(5, 11):
        ws[f"B{row}"].fill = PatternFill("solid", fgColor=FORMULA_GREEN)
        ws[f"B{row}"].protection = Protection(locked=True)
    for row in (6, 7, 8, 9):
        ws[f"B{row}"].number_format = MONEY_FORMAT
    ws["B10"].number_format = DATE_FORMAT

    ws["A12"], ws["B12"] = "事件页码", 1
    ws["C12"], ws["D12"] = "缴税页码", 1
    ws["E12"], ws["F12"] = "批次页码", 1
    for cell in ("B12", "D12", "F12"):
        ws[cell].fill = PatternFill("solid", fgColor=INPUT_BLUE)
        ws[cell].protection = Protection(locked=False)
    ws["B13"] = '="事件总数："&COUNTIF(\'激励事件明细\'!$C$2:$C$501,$B$3)&"；总页数："&MAX(1,ROUNDUP(COUNTIF(\'激励事件明细\'!$C$2:$C$501,$B$3)/25,0))'
    ws["D13"] = '="缴税总数："&COUNTIF(\'分期缴税台账\'!$A$2:$A$501,$B$3)&"；总页数："&MAX(1,ROUNDUP(COUNTIF(\'分期缴税台账\'!$A$2:$A$501,$B$3)/25,0))'
    ws["F13"] = '="批次总数："&COUNTIF(\'激励事件明细\'!$C$2:$C$501,$B$3)&"；总页数："&MAX(1,ROUNDUP(COUNTIF(\'激励事件明细\'!$C$2:$C$501,$B$3)/25,0))'
    ws["B14"] = '=IF(OR(B12<1,B12>MAX(1,ROUNDUP(COUNTIF(\'激励事件明细\'!$C$2:$C$501,$B$3)/25,0))),"页码越界","正常")'
    ws["D14"] = '=IF(OR(D12<1,D12>MAX(1,ROUNDUP(COUNTIF(\'分期缴税台账\'!$A$2:$A$501,$B$3)/25,0))),"页码越界","正常")'
    ws["F14"] = '=IF(OR(F12<1,F12>MAX(1,ROUNDUP(COUNTIF(\'激励事件明细\'!$C$2:$C$501,$B$3)/25,0))),"页码越界","正常")'

    ws["A16"] = "事件明细"
    event_display_headers = ["事件编号", "事件日期", "权益类型", "本次所得", "新增税额", "批次截止日", "校验"]
    for column, header in enumerate(event_display_headers, start=1):
        ws.cell(17, column, header)
    for display_row in range(18, 43):
        sequence = f'(($B$12-1)*25+ROW()-17)'
        match = f'MATCH($B$3&"|"&{sequence},\'激励事件明细\'!$AH$2:$AH$501,0)'
        source_columns = ("A", "G", "F", "T", "AA", "AC", "AD")
        for column, source_column in enumerate(source_columns, start=1):
            ws.cell(
                display_row,
                column,
                f'=IFERROR(INDEX(\'激励事件明细\'!${source_column}$2:${source_column}$501,{match}),"")',
            )
        ws.cell(display_row, 2).number_format = DATE_FORMAT
        for column in (4, 5):
            ws.cell(display_row, column).number_format = MONEY_FORMAT
        ws.cell(display_row, 6).number_format = DATE_FORMAT

    ws["A44"] = "分期记录"
    payment_headers = ["事件编号", "税额批次", "缴税日期", "实际金额", "批次累计已缴", "批次剩余", "校验"]
    for column, header in enumerate(payment_headers, start=1):
        ws.cell(45, column, header)
    for display_row in range(46, 71):
        sequence = f'(($D$12-1)*25+ROW()-45)'
        match = f'MATCH($B$3&"|"&{sequence},\'分期缴税台账\'!$W$2:$W$501,0)'
        source_columns = ("T", "U", "I", "K", "M", "N", "R")
        for column, source_column in enumerate(source_columns, start=1):
            ws.cell(
                display_row,
                column,
                f'=IFERROR(INDEX(\'分期缴税台账\'!${source_column}$2:${source_column}$501,{match}),"")',
            )
        ws.cell(display_row, 3).number_format = DATE_FORMAT
        for column in (4, 5, 6):
            ws.cell(display_row, column).number_format = MONEY_FORMAT

    ws["A72"] = "批次余额"
    balance_headers = ["事件编号", "税额批次", "新增税额", "有效已缴", "剩余税额", "截止日", "到期/逾期提示"]
    for column, header in enumerate(balance_headers, start=1):
        ws.cell(73, column, header)
    for display_row in range(74, 99):
        sequence = f'(($F$12-1)*25+ROW()-73)'
        match = f'MATCH($B$3&"|"&{sequence},\'激励事件明细\'!$AH$2:$AH$501,0)'
        ws.cell(display_row, 1, f'=IFERROR(INDEX(\'激励事件明细\'!$A$2:$A$501,{match}),"")')
        ws.cell(display_row, 2, f'=IFERROR(INDEX(\'激励事件明细\'!$B$2:$B$501,{match}),"")')
        ws.cell(display_row, 3, f'=IFERROR(INDEX(\'激励事件明细\'!$AA$2:$AA$501,{match}),"")')
        ws.cell(display_row, 4, f'=IF(A{display_row}="","",SUMIFS(\'分期缴税台账\'!$K$2:$K$501,\'分期缴税台账\'!$T$2:$T$501,A{display_row},\'分期缴税台账\'!$Q$2:$Q$501,"有效"))')
        ws.cell(display_row, 5, f'=IF(C{display_row}="","",MAX(0,C{display_row}-D{display_row}))')
        ws.cell(display_row, 6, f'=IFERROR(INDEX(\'激励事件明细\'!$AC$2:$AC$501,{match}),"")')
        ws.cell(display_row, 7, f'=IF(A{display_row}="","",IF(E{display_row}=0,"已缴清",IF(F{display_row}<TODAY(),"逾期","未到期")))')
        for column in (3, 4, 5):
            ws.cell(display_row, column).number_format = MONEY_FORMAT
        ws.cell(display_row, 6).number_format = DATE_FORMAT

    ws["A100"] = "到期/逾期提示"
    ws["B100"] = '=IF(COUNTIF(G74:G98,"逾期")>0,"存在逾期批次",IF(COUNTIF(G74:G98,"未到期")>0,"存在未缴清批次","本页全部已缴清"))'
    for heading_row in (16, 44, 72, 100):
        ws.cell(heading_row, 1).font = Font(bold=True, color=NAVY)
    for header_row in (17, 45, 73):
        for cell in ws[header_row][:7]:
            cell.fill = PatternFill("solid", fgColor=NAVY)
            cell.font = Font(color=WHITE, bold=True)
            cell.alignment = Alignment(horizontal="center")
    add_list_validation(ws, "=员工主数据编号", ["B3"])
    page_validation = DataValidation(
        type="whole",
        operator="between",
        formula1="1",
        formula2="20",
        allow_blank=False,
    )
    ws.add_data_validation(page_validation)
    for cell in ("B12", "D12", "F12"):
        page_validation.add(cell)
    set_column_widths(ws, {1: 22, 2: 22, 3: 18, 4: 18, 5: 18, 6: 18, 7: 28})
    ws.print_area = "A1:G100"
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = "A17:G42"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    protect(ws)


def create_policy_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("政策及公告")
    headers = ["类别", "文件名称", "发布日期", "文号/公告编号", "采用内容", "来源链接"]
    ws.append(headers)
    rows = [
        ("公司公告", "关于向2023年股票期权与限制性股票激励计划激励对象授予权益的公告", date(2024, 7, 2), "2024-043", "实际授予862.50万份/股、623人；授予日2024-07-01", "https://disc.static.szse.cn/disc/disk03/finalpage/2024-07-01/d4e22bf2-dc60-41be-ab59-090c05df4556.PDF"),
        ("公司公告", "股票期权授予登记完成公告", date(2024, 7, 25), "2024-045", "登记802万份、557人；登记完成日2024-07-24", "https://static.cninfo.com.cn/finalpage/2024-07-25/1220719491.PDF"),
        ("公司公告", "限制性股票授予登记完成公告", date(2024, 8, 26), "2024-054", "登记802万股、557人；上市及登记日2024-08-27", "https://static.cninfo.com.cn/finalpage/2024-08-26/1220961881.PDF"),
        ("行情", "深圳证券交易所股票历史行情", date(2024, 8, 27), "无", "002967当日今收13.06元", "https://www.szse.cn/api/report/ShowReport/data?SHOWTYPE=JSON&CATALOGID=1815_stock&TABKEY=tab1&txtDMorJC=002967&txtBeginDate=2024-08-27"),
        ("税收政策", "关于延续实施上市公司股权激励有关个人所得税政策的公告", date(2023, 8, 18), "财政部 税务总局公告2023年第25号", "同年度两次以上股权激励合并、单独计税", "https://fgk.chinatax.gov.cn/zcfgk/c102416/c5211082/content.html"),
        ("税收政策", "关于上市公司股权激励有关个人所得税政策的公告", date(2024, 4, 17), "财政部 税务总局公告2024年第2号", "备案后自事件日起不超过36个月缴税；离职前缴清", "https://jiangsu.chinatax.gov.cn/art/2024/4/28/art_8349_458449.html"),
        ("税收政策", "关于股权激励有关个人所得税问题的通知", date(2009, 8, 24), "国税函〔2009〕461号", "限制性股票所得公式及个人口径分母", "https://fgk.chinatax.gov.cn/zcfgk/c100012/c5194079/content.html"),
        ("税收政策", "关于个人股票期权所得征收个人所得税问题的通知", date(2005, 3, 28), "财税〔2005〕35号", "股票期权所得公式；市场价为行权日收盘价", "https://fgk.chinatax.gov.cn/zcfgk/c102416/c5203004/content.html"),
        ("公司公告", "2024年中期权益分派实施公告", date(2025, 1, 6), "2025-002", "2025-01-10起行权价14.32", "https://static.cninfo.com.cn/finalpage/2025-01-06/1222225049.PDF"),
        ("公司公告", "2024年年度权益分派实施公告", date(2025, 4, 23), "2025-031", "2025-04-29起行权价14.18", "https://static.cninfo.com.cn/finalpage/2025-04-23/1223212213.PDF"),
        ("公司公告", "2025年中期权益分派实施公告", date(2025, 9, 19), "2025-060", "2025-09-26起行权价14.04", "https://static.cninfo.com.cn/finalpage/2025-09-19/1224668254.PDF"),
    ]
    for row in rows:
        ws.append(row)
    style_header(ws)
    fill_cells(ws, set(range(1, 7)), 2, ws.max_row, PARAM_GRAY, True)
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 3).number_format = DATE_FORMAT
        ws.cell(row, 6).hyperlink = ws.cell(row, 6).value
        ws.cell(row, 6).style = "Hyperlink"
    add_table(ws, "tblPolicies", ws.max_row, len(headers))
    set_print(ws, ws.max_row, len(headers))
    set_column_widths(ws, {1: 14, 2: 48, 3: 14, 4: 30, 5: 52, 6: 55})
    protect(ws)


def define_names(wb: Workbook) -> None:
    names = {
        "员工主数据编号": "'计划参数'!$I$2:$I$501",
        "员工编号列表": "'计划参数'!$I$2:$I$501",
        "事件编号列表": "'激励事件明细'!$A$2:$A$501",
        "税额批次列表": "'激励事件明细'!$B$2:$B$501",
    }
    for name, reference in names.items():
        wb.defined_names.add(DefinedName(name, attr_text=reference))


def build_workbook() -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    create_instructions(wb)
    create_plan_parameters(wb)
    create_event_sheet(wb)
    create_summary(wb)
    create_ledger(wb)
    create_person_sheet(wb)
    create_tax_rules(wb)
    create_policy_sheet(wb)
    define_names(wb)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    return OUTPUT


if __name__ == "__main__":
    result = build_workbook()
    print(result)
